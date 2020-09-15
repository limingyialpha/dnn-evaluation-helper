# This import fixes the problem that specifying the type of an object
# in its module definition raises error
from __future__ import annotations

from importlib.resources import path
from pathlib import Path
from typing import Generator, Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class ExcelSheet:
    """Represents an excel worksheet

    This class encapsulates all the attributes and behaviours of
    an excel worksheet.
    It hides, regulates and standardizes the access to other excel
    processing libraries, thus protecting developers from making
    mistakes and simplifies the excel processing in the project.
    It is encouraged to use this class to represent all the excel
    worksheets in the project.
    Please notice that the excel worksheet entity behind this class
    is always the active worksheet of an excel workbook.
    """

    def __init__(self, workbook: Workbook):
        """Initialize an excel worksheet

        It is not encouraged to use the constructor directly to construct new objects.
        Instead, use the following factory methods:
        -- open_resource:
            opens an excel worksheet from the project's resources
        -- create
            create a new excel worksheet from nothing

        Parameters
        ----------
        workbook:
            an opened excel workbook from a 3rd party library
        """
        self._workbook = workbook
        self._worksheet = workbook.active

    @classmethod
    def open_resource(cls, package: str, resource: str) -> ExcelSheet:
        """Opens an image from the project's resources

        Factory method to initialize an object

        Parameters
        ----------
        package:
            package name, starting from the top level package to the package
            containing the resource, using . as separator
            For example: 'app.utilities'
        resource
            resource file name with extension
            For example: 'i_hate_programming.txt'

        Returns
        -------
        excel worksheet:
            an opened excel worksheet object
        """
        with path(package, resource) as p:
            wb = load_workbook(p)
        return cls(wb)

    @classmethod
    def create(cls) -> ExcelSheet:
        """Creates an excel worksheet from nothing

        Factory method to initialize an object

        Returns
        -------
        excel worksheet:
            an opened new and empty excel worksheet object
        """
        wb = Workbook()
        return cls(wb)

    def set_cell(self, row: int, column: int, value: Any):
        """Sets the value of a cell

        Parameters
        ----------
        row:
            the row number starting with index 1
        column:
            the column number starting with index 1
        value:
            the value of the cell
        """
        self._worksheet.cell(row, column, value)

    def iter_row_values(self) -> Generator[Any]:
        """Iterate through the actual row values

        The actual row values does not contain the first row and the first column
        Because they are responsible for the semantics and value type of each rows
        and columns

        Returns
        -------
        generator:
            a generator. when iterating, it returns a tuple with values in each row.
        """
        return self._worksheet.iter_rows(
            min_row=2, max_row=self._worksheet.max_row,
            min_col=2, max_col=self._worksheet.max_column,
            values_only=True)

    def save(self, save_path: Path):
        """Saves the excel worksheet

        The excel workbook that contains the worksheet will be saved
        to the path specified by the given path object.
        if the given path does not exist,
        this path will be automatically generated

        Parameters
        ----------
        save_path:
            the path object specifying the saving location
        """
        save_path.parent.mkdir(parents=True, exist_ok=True)
        self._workbook.save(save_path)
